import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment

save_path = '03. 엑셀자동화/total.xlsx'

# 기존의 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path)

# 시트 선택
ws = wb['data']

# 열 너비
ws.column_dimensions['B'].width = 20

# 행 높이 변경
ws.row_dimensions[1].height = 30

# 폰트 변경
ft = Font(size=12, color='0077ff', bold=True, name='돋움')
ws['A1'].font = ft

# 경계선 지정하기
border_type1 = Side(color='000000', border_style='thin')
border_type2 = Side(color='000000', border_style='thick')

#아래 경계선
border_bottom = Border(bottom=border_type1)
ws['A4'].border = border_bottom

# 모든 방향 경계선
border_all = Border(left=border_type2, right=border_type2, 
                    bottom=border_type2, top=border_type2)
ws['B4'].border = border_all

# 정렬하기
alignment_type = Alignment(horizontal='center', vertical='center')

for row in ws.iter_rows(max_row=1):
    for cell in row:
        cell.alignment = alignment_type
        cell.border = border_bottom
        cell.font = ft

# 저장
wb.save(save_path)