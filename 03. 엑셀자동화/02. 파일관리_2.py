import openpyxl


# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()


# 새로운 시트 생성
ws = wb.create_sheet('2030.01')

# 모든 시트 이름 출력
print(wb.sheetnames)

# 시트 삭제
del wb['Sheet']

# 엑셀 저장
wb.save('C:\Users\jyhuh\Desktop\Dev\automation\03. 엑셀자동화\거래처A매입현황.xlsx')