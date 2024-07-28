import openpyxl

# 엑셀 파일 불러오기
wb = openpyxl.load_workbook('03. 엑셀자동화/total.xlsx')

# 현재 활성화 된 시트 선택
ws = wb.active

# 제품이름 리스트
name_list = []

for row in ws.iter_rows(min_row=2, min_col=2):
    name = row[0].value
    if name not in name_list:   # name이 포함되어 있지 않다면
        name_list.append(name)
        wb.create_sheet(name) # sheet이름 생성
    data = []
    for cell in row:
        data.append(cell.value)
    wb[name].append(data)

wb.save('03. 엑셀자동화/total.xlsx')